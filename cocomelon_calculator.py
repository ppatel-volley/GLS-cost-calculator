#!/usr/bin/env python3
"""
CoComelon AWS GameLift Streams Cost Calculator

Interactive spreadsheet generator for analyzing streaming costs across
different GPU instance types with adjustable DAU and usage patterns.

Key Features:
- DAU-driven capacity scaling
- Time zone-aware usage modeling (East to West coast)
- AlwaysOn vs OnDemand cost comparison
- Interactive Excel parameters for scenario modeling
- Comprehensive cost analysis (day/week/month/year)
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CoComelon Game Configuration
GAME_CONFIG = {
    'target_dau': 55000,
    'session_duration_min': 15,
    'session_duration_max': 60,
    'peak_usage_hours': {
        'weekday_start': 7,   # 7am PT (brief morning window)
        'weekday_end': 19,    # 7pm PT  
        'weekend_start': 7,   # 7am PT (kids wake up on weekends)
        'weekend_end': 19     # 7pm PT
    },
    'concurrent_ratio': 0.0064,  # 350 concurrent / 55000 DAU = ~0.64%
    'weekend_multiplier': 0.8,   # 80% of weekday peak
    'off_peak_multiplier': 0.1   # 10% of peak during off hours
}

# AWS GameLift Streams Instance Configuration
INSTANCE_TYPES = {
    'gen4n_high': {
        'description': 'NVIDIA T4 - Cost optimized',
        'gpu': 'NVIDIA T4 Tensor GPU',
        'vcpu': 4,
        'vram_gb': 8,
        'ram_gb': 16,
        'sessions_per_host': 2,
        'hourly_rate': 0.50,
        'notes': 'Most cost-effective, sufficient for 1080p kids content'
    },
    'gen5n_high': {
        'description': 'NVIDIA A10G - Performance balanced', 
        'gpu': 'NVIDIA A10G Tensor GPU',
        'vcpu': 4,
        'vram_gb': 12,
        'ram_gb': 16,
        'sessions_per_host': 4,
        'hourly_rate': 0.77,
        'notes': 'Better VRAM, 4:1 multi-tenancy efficiency'
    },
    'g6n_high': {
        'description': 'NVIDIA L4 - High capacity (Early Access)',
        'gpu': 'NVIDIA L4 Tensor GPU', 
        'vcpu': 8,
        'vram_gb': 24,
        'ram_gb': 32,
        'sessions_per_host': 8,
        'hourly_rate': 1.50,
        'notes': 'Maximum efficiency with 8:1 multi-tenancy'
    }
}

# Cost Model Configuration
COST_MODEL = {
    'always_on_percentage': 0.30,  # 30% always-on capacity
    'storage_cost_per_gb_month': 0.03,  # GameLift storage cost
    'storage_gb_required': 100,  # Estimated game asset size
    'subscription_price': 12.99  # Monthly subscription revenue per user
}

def calculate_concurrent_streams(dau, concurrent_ratio=GAME_CONFIG['concurrent_ratio']):
    """Calculate peak concurrent streams from DAU."""
    return int(dau * concurrent_ratio)

def get_hourly_usage_pattern():
    """Create 24-hour usage pattern for weekdays and weekends with realistic children's viewing habits."""
    patterns = {}
    
    # Calculate multiplier for exactly 2 users during quiet hours (9pm-5am)
    # 2 users out of 55,000 DAU = 2/55000 = 0.0000364
    # But we need this as a multiplier of peak concurrent (352), so: 2/352 = 0.00568
    quiet_hour_multiplier = 2 / 352  # ~0.00568 for exactly 2 concurrent users
    
    # Weekday pattern: brief morning window (7-8am), minimal during school, peak after school (5-7pm)
    weekday_multipliers = [
        quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier, 0.01, 0.05, 0.25,  # 12am-5am: 2 users, 6am: 2-3, 7am morning rush
        0.03, 0.01, 0.01, 0.01, 0.01, 0.01, 0.01, 0.02,       # 8am-3pm: kids in school - very minimal
        0.05, 1.00, 1.00, 0.60, 0.10, quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier      # 4pm: pickup, 5-6pm: PEAK, 9pm-11pm: 2 users
    ]
    
    # Weekend pattern: gradual morning start (7am), flexible usage all day (7am-7pm)
    weekend_multipliers = [
        quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier, 0.002, 0.01, 0.15,  # 12am-5am: 2 users, 6am: minimal, 7am: wake up
        0.35, 0.50, 0.60, 0.65, 0.70, 0.75, 0.70, 0.65,      # 8am-3pm: steady weekend usage
        0.70, 0.75, 0.70, 0.55, 0.25, quiet_hour_multiplier, quiet_hour_multiplier, quiet_hour_multiplier      # 4pm-8pm: evening peak, 9pm-11pm: 2 users
    ]
    
    patterns['weekday'] = weekday_multipliers
    patterns['weekend'] = weekend_multipliers
    
    return patterns

def create_excel_calculator():
    """Create interactive Excel calculator with formula-driven calculations."""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default worksheet
    wb.remove(wb.active)
    
    # Create Inputs sheet
    inputs_ws = wb.create_sheet('Inputs_Controls')
    
    # Set up interactive parameters
    inputs_data = [
        ['COCOMELON GAMELIFT STREAMS CALCULATOR', ''],
        ['', ''],
        ['GAME CONFIGURATION', ''],
        ['Daily Active Users (DAU)', GAME_CONFIG['target_dau']],
        ['Peak Concurrent Streams', '=ROUND(B4*0.0064,0)'],
        ['Session Duration (min)', f"{GAME_CONFIG['session_duration_min']}-{GAME_CONFIG['session_duration_max']}"],
        ['Concurrent Stream % of DAU', '=ROUND((B5/B4)*100,2)'],
        ['', ''],
        ['USAGE PATTERNS', ''],
        ['Weekday Available Hours', '7:00-19:00 PT (peak after school 17:00-18:00)'],
        ['Weekend Available Hours', '7:00-19:00 PT (flexible usage all day)'],
        ['Weekend Usage Multiplier', GAME_CONFIG['weekend_multiplier']],
        ['Off-Peak Usage Multiplier', GAME_CONFIG['off_peak_multiplier']],
        ['', ''],
        ['COST PARAMETERS', ''],
        ['Always-On Capacity %', COST_MODEL['always_on_percentage']],
        ['Storage Cost ($/GB/month)', COST_MODEL['storage_cost_per_gb_month']],
        ['Storage Required (GB)', COST_MODEL['storage_gb_required']],
        ['Monthly Subscription Revenue', COST_MODEL['subscription_price']],
        ['', ''],
        ['INSTANCE PRICING ($/hour)', ''],
        ['gen4n_high (T4, 2 sessions)', INSTANCE_TYPES['gen4n_high']['hourly_rate']],
        ['gen5n_high (A10G, 4 sessions)', INSTANCE_TYPES['gen5n_high']['hourly_rate']],
        ['g6n_high (L4, 8 sessions)', INSTANCE_TYPES['g6n_high']['hourly_rate']],
    ]
    
    for row_idx, (param, value) in enumerate(inputs_data, 1):
        inputs_ws.cell(row=row_idx, column=1, value=param)
        if isinstance(value, str) and value.startswith('='):
            inputs_ws.cell(row=row_idx, column=2).value = value
        else:
            inputs_ws.cell(row=row_idx, column=2, value=value)
    
    # Format the inputs sheet
    for row in inputs_ws.iter_rows():
        for cell in row:
            if cell.column == 1:  # Parameter column
                if cell.value and cell.value in ['GAME CONFIGURATION', 'USAGE PATTERNS', 'COST PARAMETERS', 'INSTANCE PRICING ($/hour)']:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            elif cell.column == 2:  # Value column
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, str):
                    cell.number_format = '#,##0.00'
    
    # Set column widths
    inputs_ws.column_dimensions['A'].width = 30
    inputs_ws.column_dimensions['B'].width = 15
    
    # Create Calculations sheet
    calc_ws = wb.create_sheet('Calculations')
    
    # Set up calculation headers with helper columns
    calc_headers = [
        ['Instance Type', 'Hourly Rate', 'Sessions/Host', 'Peak Hosts Needed', 
         'Always-On Hosts', 'Max On-Demand', 'WD Always-On Cost', 'WD OnDemand Cost',
         'WE Always-On Cost', 'WE OnDemand Cost', 'Weekday Daily Cost', 
         'Weekend Daily Cost', 'Weekly Cost', 'Monthly Cost', 'Annual Cost',
         'Cost per User/Month', '% of $12.99 Sub']
    ]
    
    for row_idx, row_data in enumerate(calc_headers, 1):
        for col_idx, header in enumerate(row_data, 1):
            cell = calc_ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
    
    # Add instance calculations with formulas
    instances = [
        ('gen4n_high', 22, 2, 'T4'),
        ('gen5n_high', 23, 4, 'A10G'), 
        ('g6n_high', 24, 8, 'L4')
    ]
    
    for row_idx, (instance_name, rate_ref, sessions_per_host, gpu) in enumerate(instances, 2):
        # Instance name
        calc_ws.cell(row=row_idx, column=1, value=f"{instance_name} ({gpu})")
        
        # Hourly rate (reference to inputs)
        calc_ws.cell(row=row_idx, column=2, value=f"=Inputs_Controls!B{rate_ref}")
        
        # Sessions per host
        calc_ws.cell(row=row_idx, column=3, value=sessions_per_host)
        
        # Peak hosts needed = ROUNDUP(Peak_Concurrent_Streams / Sessions_Per_Host, 0)
        calc_ws.cell(row=row_idx, column=4, 
                    value=f"=ROUNDUP(Inputs_Controls!B5/C{row_idx}, 0)")
        
        # Always-on hosts = Peak_Hosts * Always_On_Percentage  
        calc_ws.cell(row=row_idx, column=5,
                    value=f"=ROUNDUP(D{row_idx}*Inputs_Controls!B16, 0)")
        
        # Max on-demand hosts
        calc_ws.cell(row=row_idx, column=6, value=f"=D{row_idx}-E{row_idx}")
        
        # Break down cost calculation into manageable chunks
        patterns = get_hourly_usage_pattern()
        
        # Calculate weighted averages for simplified formulas
        weekday_avg = sum(patterns['weekday']) / 24
        weekend_avg = sum(patterns['weekend']) / 24
        
        # Peak hour multipliers for OnDemand calculation
        weekday_peak = max(patterns['weekday'])
        weekend_peak = max(patterns['weekend'])
        
        # Column 7: Weekday Always-On Cost (24/7 constant)
        calc_ws.cell(row=row_idx, column=7, value=f"=E{row_idx}*B{row_idx}*24")
        
        # Column 8: Weekday OnDemand Cost (variable based on usage patterns)
        # Simplified: Use weighted average of additional capacity needed
        weekday_ondemand_formula = f"=MAX(0,ROUNDUP(Inputs_Controls!B5*{weekday_peak}/C{row_idx},0)-E{row_idx})*B{row_idx}*2" \
                                 + f"+MAX(0,ROUNDUP(Inputs_Controls!B5*{weekday_avg}/C{row_idx},0)-E{row_idx})*B{row_idx}*22"
        calc_ws.cell(row=row_idx, column=8, value=weekday_ondemand_formula)
        
        # Column 9: Weekend Always-On Cost (24/7 constant)
        calc_ws.cell(row=row_idx, column=9, value=f"=E{row_idx}*B{row_idx}*24")
        
        # Column 10: Weekend OnDemand Cost (variable based on usage patterns)
        weekend_ondemand_formula = f"=MAX(0,ROUNDUP(Inputs_Controls!B5*{weekend_peak}/C{row_idx},0)-E{row_idx})*B{row_idx}*8" \
                                  + f"+MAX(0,ROUNDUP(Inputs_Controls!B5*{weekend_avg}/C{row_idx},0)-E{row_idx})*B{row_idx}*16"
        calc_ws.cell(row=row_idx, column=10, value=weekend_ondemand_formula)
        
        # Column 11: Weekday Daily Cost = Always-On + OnDemand
        calc_ws.cell(row=row_idx, column=11, value=f"=G{row_idx}+H{row_idx}")
        
        # Column 12: Weekend Daily Cost = Always-On + OnDemand  
        calc_ws.cell(row=row_idx, column=12, value=f"=I{row_idx}+J{row_idx}")
        
        # Column 13: Weekly cost = weekday * 5 + weekend * 2
        calc_ws.cell(row=row_idx, column=13, value=f"=K{row_idx}*5+L{row_idx}*2")
        
        # Column 14: Monthly cost = Weekly * 4.33
        calc_ws.cell(row=row_idx, column=14, value=f"=M{row_idx}*4.33")
        
        # Column 15: Annual cost = Weekly * 52  
        calc_ws.cell(row=row_idx, column=15, value=f"=M{row_idx}*52")
        
        # Column 16: Cost per User per Month = (Monthly Cost + Storage) / DAU
        calc_ws.cell(row=row_idx, column=16, value=f"=(N{row_idx}+Inputs_Controls!B17*Inputs_Controls!B18)/Inputs_Controls!B4")
        
        # Column 17: Percentage of $12.99 subscription
        calc_ws.cell(row=row_idx, column=17, value=f"=P{row_idx}/Inputs_Controls!B19")
    
    # Format calculations sheet
    for col in range(1, 18):
        calc_ws.column_dimensions[get_column_letter(col)].width = 15
        
    # Add currency formatting to cost columns (7-16)
    for row in range(2, 5):
        for col in range(7, 17):
            calc_ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
            
    # Format percentage column (17)
    for row in range(2, 5):
        calc_ws.cell(row=row, column=17).number_format = '0.00%'
    
    # Create Summary Dashboard
    summary_ws = wb.create_sheet('Cost_Dashboard')
    
    # Create summary table
    summary_data = [
        ['COCOMELON STREAMING COST ANALYSIS', '', '', ''],
        ['', '', '', ''],
        ['Instance Comparison', 'gen4n_high', 'gen5n_high', 'g6n_high'],
        ['GPU Type', 'NVIDIA T4', 'NVIDIA A10G', 'NVIDIA L4'],
        ['VRAM (GB)', '8', '12', '24'],
        ['Sessions per Host', '2', '4', '8'],
        ['Hourly Rate', '=Calculations!B2', '=Calculations!B3', '=Calculations!B4'],
        ['Peak Hosts Required', '=Calculations!D2', '=Calculations!D3', '=Calculations!D4'],
        ['Always-On Hosts', '=Calculations!E2', '=Calculations!E3', '=Calculations!E4'],
        ['', '', '', ''],
        ['MONTHLY COSTS', '', '', ''],
        ['Streaming Cost', '=Calculations!N2', '=Calculations!N3', '=Calculations!N4'],
        ['Storage Cost', "=Inputs_Controls!B17*Inputs_Controls!B18", "=Inputs_Controls!B17*Inputs_Controls!B18", "=Inputs_Controls!B17*Inputs_Controls!B18"],
        ['Total Monthly Cost', '=B12+B13', '=C12+C13', '=D12+D13'],
        ['', '', '', ''],
        ['COST PER USER ANALYSIS', '', '', ''],
        ['Monthly Cost per DAU', '=B14/Inputs_Controls!B4', '=C14/Inputs_Controls!B4', '=D14/Inputs_Controls!B4'],
        ['% of Subscription Revenue', '=B17/Inputs_Controls!B19', '=C17/Inputs_Controls!B19', '=D17/Inputs_Controls!B19'],
        ['', '', '', ''],
        ['ANNUAL PROJECTIONS', '', '', ''],
        ['Annual Streaming Cost', '=Calculations!O2', '=Calculations!O3', '=Calculations!O4'],
        ['Annual Total Cost', '=B21+(B13*12)', '=C21+(C13*12)', '=D21+(D13*12)'],
        ['', '', '', ''],
        ['COST SAVINGS vs gen4n_high', '', '', ''],
        ['Monthly Difference', 'Baseline', '=C14-B14', '=D14-B14'],
        ['Annual Difference', 'Baseline', '=C22-B22', '=D22-B22'],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format headers
            if row_idx in [1, 3, 11, 16, 20, 24] and col_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    # Format currency cells
    currency_rows = [7, 12, 13, 14, 17, 21, 22, 25, 26]
    for row in currency_rows:
        for col in range(2, 5):
            summary_ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
    
    # Format percentage cells
    for col in range(2, 5):
        summary_ws.cell(row=18, column=col).number_format = '0.00%'
    
    # Set column widths
    for col in range(1, 5):
        summary_ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Create DAU Sensitivity Analysis sheet
    sensitivity_ws = wb.create_sheet('DAU_Sensitivity')
    
    # Create DAU scenario table
    dau_scenarios = [25000, 35000, 45000, 55000, 65000, 75000, 100000]
    
    sensitivity_headers = ['DAU Scenario'] + [f'{dau:,}' for dau in dau_scenarios]
    for col_idx, header in enumerate(sensitivity_headers, 1):
        cell = sensitivity_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
    
    # Add calculations for different DAU levels
    metrics = [
        'Peak Concurrent Streams',
        'gen4n_high Monthly Cost',  
        'gen5n_high Monthly Cost',
        'g6n_high Monthly Cost',
        'Cost per User (gen4n_high)',
        'Cost per User (gen5n_high)', 
        'Cost per User (g6n_high)'
    ]
    
    for row_idx, metric in enumerate(metrics, 2):
        sensitivity_ws.cell(row=row_idx, column=1, value=metric)
        
        for col_idx, dau in enumerate(dau_scenarios, 2):
            concurrent = int(dau * GAME_CONFIG['concurrent_ratio'])
            
            if metric == 'Peak Concurrent Streams':
                value = concurrent
            elif 'Monthly Cost' in metric:
                # Simplified cost calculation for sensitivity
                instance_type = metric.split()[0]
                sessions = INSTANCE_TYPES[instance_type]['sessions_per_host']
                rate = INSTANCE_TYPES[instance_type]['hourly_rate']
                hosts_needed = np.ceil(concurrent / sessions)
                always_on = int(hosts_needed * COST_MODEL['always_on_percentage'])
                
                # Simplified daily cost estimation
                daily_cost = (3 * hosts_needed + 21 * always_on) * rate  # Weekday
                weekend_cost = (15 * int(concurrent * 0.8 / sessions) + 9 * always_on) * rate
                monthly_cost = (daily_cost * 5 + weekend_cost * 2) * 4.33
                value = f"=ROUND({monthly_cost}, 0)"
                
            elif 'Cost per User' in metric:
                instance_type = metric.split()[-1].strip('()')
                value = f"=INDIRECT(ADDRESS(ROW()-3, COLUMN()))/INDIRECT(ADDRESS(1, COLUMN()))"
            
            cell = sensitivity_ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, str) and not value.startswith('='):
                try:
                    cell.value = float(value.replace(',', ''))
                except:
                    pass
    
    # Format sensitivity sheet
    for col in range(1, len(dau_scenarios) + 2):
        sensitivity_ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add formatting for currency values
    for row in range(3, 5):  # Monthly cost rows
        for col in range(2, len(dau_scenarios) + 2):
            sensitivity_ws.cell(row=row, column=col).number_format = '"$"#,##0'
    
    for row in range(5, 8):  # Cost per user rows  
        for col in range(2, len(dau_scenarios) + 2):
            sensitivity_ws.cell(row=row, column=col).number_format = '"$"0.0000'
    
    # Create Usage Pattern Analysis sheet
    usage_ws = wb.create_sheet('Usage_Patterns')
    
    # Create 24-hour usage pattern table
    usage_headers = ['Hour', 'Time', 'WD Total Hourly Active Users', 'WE Total Hourly Active Users', 
                    'WD Concurrent Streaming', 'WE Concurrent Streaming', 'WD Hosts Needed (gen5n)', 
                    'WE Hosts Needed (gen5n)']
    
    # Add explanatory note
    usage_ws.cell(row=1, column=1, value="USAGE PATTERNS - 24 Hour Analysis")
    usage_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # Add column explanations
    explanations = [
        "Column Definitions:",
        "• Total Hourly Active Users: Users nationwide opening the app during this hour (across all US time zones)", 
        "• Concurrent Streaming: Users simultaneously streaming at peak of this hour",
        "• Based on children's viewing habits (school schedules, bedtime, etc.)"
    ]
    
    for i, explanation in enumerate(explanations):
        usage_ws.cell(row=2+i, column=1, value=explanation)
        if i == 0:
            usage_ws.cell(row=2+i, column=1).font = Font(bold=True)
    
    # Headers start at row 6
    header_row = 6
    for col_idx, header in enumerate(usage_headers, 1):
        cell = usage_ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
    
    # Get usage patterns
    hourly_patterns = get_hourly_usage_pattern()
    
    # Calculate active users and concurrent streams for each hour
    for hour in range(24):
        row_idx = hour + 7  # Data starts at row 7 (header at row 6)
        
        # Hour and time display
        usage_ws.cell(row=row_idx, column=1, value=hour)
        time_display = f"{hour:02d}:00"
        if hour < 12:
            time_display += " AM"
        elif hour == 12:
            time_display += " PM"
        else:
            time_display = f"{hour-12:02d}:00 PM"
        usage_ws.cell(row=row_idx, column=2, value=time_display)
        
        # Calculate active users (assuming active users follow same pattern as concurrent)
        weekday_multiplier = hourly_patterns['weekday'][hour]
        weekend_multiplier = hourly_patterns['weekend'][hour]
        
        # Total active users nationwide = DAU * hourly usage multiplier * activity factor
        # This represents total users who open the app/are active during this hour across all time zones
        weekday_active = f"=ROUND(Inputs_Controls!B4*{weekday_multiplier}*0.2,0)"  # 20% activity factor during peak patterns
        weekend_active = f"=ROUND(Inputs_Controls!B4*{weekend_multiplier}*0.2,0)"
        
        usage_ws.cell(row=row_idx, column=3, value=weekday_active)
        usage_ws.cell(row=row_idx, column=4, value=weekend_active)
        
        # Concurrent streams
        weekday_concurrent = f"=ROUND(Inputs_Controls!B5*{weekday_multiplier},0)"
        weekend_concurrent = f"=ROUND(Inputs_Controls!B5*{weekend_multiplier},0)"
        
        usage_ws.cell(row=row_idx, column=5, value=weekday_concurrent)
        usage_ws.cell(row=row_idx, column=6, value=weekend_concurrent)
        
        # Hosts needed (using gen5n_high as example - 4 sessions per host)
        weekday_hosts = f"=ROUNDUP(E{row_idx}/4,0)"
        weekend_hosts = f"=ROUNDUP(F{row_idx}/4,0)"
        
        usage_ws.cell(row=row_idx, column=7, value=weekday_hosts)
        usage_ws.cell(row=row_idx, column=8, value=weekend_hosts)
    
    # Format the usage patterns sheet
    for col in range(1, 9):
        usage_ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Add number formatting
    for row in range(7, 31):  # All data rows (now start at row 7)
        for col in range(3, 9):  # User count columns
            usage_ws.cell(row=row, column=col).number_format = '#,##0'
    
    # Add summary section
    summary_start_row = 33  # Moved down due to explanatory text
    usage_ws.cell(row=summary_start_row, column=1, value="USAGE PATTERN SUMMARY").font = Font(bold=True, size=12)
    
    summary_data = [
        ['Peak Weekday Hour', f"=INDEX(B7:B30,MATCH(MAX(E7:E30),E7:E30,0))"],
        ['Peak Weekend Hour', f"=INDEX(B7:B30,MATCH(MAX(F7:F30),F7:F30,0))"],
        ['Peak Weekday Concurrent', '=MAX(E7:E30)'],
        ['Peak Weekend Concurrent', '=MAX(F7:F30)'],
        ['Min Weekday Concurrent', '=MIN(E7:E30)'],
        ['Min Weekend Concurrent', '=MIN(F7:F30)'],
        ['Weekday Usage Ratio (Peak/Min)', f'=B{summary_start_row + 3}/B{summary_start_row + 5}'],
        ['Weekend Usage Ratio (Peak/Min)', f'=B{summary_start_row + 4}/B{summary_start_row + 6}'],
        ['', ''],
        ['QUIET HOURS (9PM-5AM)', ''],
        ['Expected Concurrent Users', '2'],
        ['Actual Formula Result', '=ROUND(Inputs_Controls!B5*{:.6f},0)'.format(2/352)],
    ]
    
    for idx, (label, formula) in enumerate(summary_data):
        row = summary_start_row + 1 + idx
        usage_ws.cell(row=row, column=1, value=label)
        usage_ws.cell(row=row, column=2, value=formula)
        if 'Concurrent' in label:
            usage_ws.cell(row=row, column=2).number_format = '#,##0'
        elif 'Ratio' in label:
            usage_ws.cell(row=row, column=2).number_format = '0.0"x"'
    
    # Save the workbook
    filename = f'CoComelon_GameLift_Calculator_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(filename)
    
    return filename

def generate_analysis_report():
    """Generate text analysis report."""
    
    report_lines = [
        "COCOMELON GAMELIFT STREAMS - COST ANALYSIS REPORT",
        "=" * 60,
        "",
        "EXECUTIVE SUMMARY:",
        f"• Target DAU: {GAME_CONFIG['target_dau']:,} users",
        f"• Peak Concurrent: {calculate_concurrent_streams(GAME_CONFIG['target_dau']):,} streams", 
        "• Usage Pattern: East to West coast (children's viewing hours)",
        f"• Session Length: {GAME_CONFIG['session_duration_min']}-{GAME_CONFIG['session_duration_max']} minutes",
        "",
        "INSTANCE TYPE COMPARISON:",
        ""
    ]
    
    # Calculate costs for each instance type
    concurrent_peak = calculate_concurrent_streams(GAME_CONFIG['target_dau'])
    
    for name, config in INSTANCE_TYPES.items():
        hosts_needed = np.ceil(concurrent_peak / config['sessions_per_host'])
        always_on_hosts = int(hosts_needed * COST_MODEL['always_on_percentage'])
        
        # Simplified cost calculation 
        daily_weekday = (3 * hosts_needed + 21 * always_on_hosts) * config['hourly_rate']
        daily_weekend = (15 * int(concurrent_peak * 0.8 / config['sessions_per_host']) + 9 * always_on_hosts) * config['hourly_rate']
        monthly_streaming = (daily_weekday * 5 + daily_weekend * 2) * 4.33
        monthly_storage = COST_MODEL['storage_cost_per_gb_month'] * COST_MODEL['storage_gb_required']
        monthly_total = monthly_streaming + monthly_storage
        
        cost_per_user = monthly_total / GAME_CONFIG['target_dau']
        revenue_percentage = (cost_per_user / COST_MODEL['subscription_price']) * 100
        
        report_lines.extend([
            f"{name.upper()} ({config['description']}):",
            f"  • GPU: {config['gpu']}",
            f"  • Sessions per host: {config['sessions_per_host']}",  
            f"  • Hosts required: {int(hosts_needed)}",
            f"  • Monthly cost: ${monthly_total:,.2f}",
            f"  • Cost per user: ${cost_per_user:.4f} ({revenue_percentage:.1f}% of subscription)",
            ""
        ])
    
    report_lines.extend([
        "KEY FINDINGS:",
        "• Multi-tenancy is critical for cost efficiency",
        "• g6n_high offers best cost-per-stream with 6:1 sessions",  
        "• All options are under 5% of subscription revenue",
        "• AlwaysOn + OnDemand hybrid reduces costs significantly",
        "",
        "RECOMMENDATION:",
        "Start with gen4n_high for initial launch, evaluate g6n_high",
        "for scale efficiency once early access is available.",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ])
    
    report_content = "\n".join(report_lines)
    
    # Save report
    report_filename = f'CoComelon_Analysis_Report_{datetime.now().strftime("%Y%m%d")}.txt'
    with open(report_filename, 'w') as f:
        f.write(report_content)
    
    return report_filename, report_content

if __name__ == "__main__":
    print("CoComelon GameLift Streams Cost Calculator")
    print("=" * 50)
    
    # Generate Excel calculator
    print("\nCreating interactive Excel calculator...")
    excel_file = create_excel_calculator()
    print(f"Excel calculator saved: {excel_file}")
    
    # Generate analysis report
    print("\nGenerating analysis report...")
    report_file, report_content = generate_analysis_report()
    print(f"Analysis report saved: {report_file}")
    
    # Print summary to console
    print("\n" + "=" * 50)
    print("QUICK SUMMARY:")
    print("=" * 50)
    concurrent = calculate_concurrent_streams(GAME_CONFIG['target_dau'])
    print(f"Target DAU: {GAME_CONFIG['target_dau']:,}")
    print(f"Peak Concurrent: {concurrent:,}")
    
    for name, config in INSTANCE_TYPES.items():
        hosts = np.ceil(concurrent / config['sessions_per_host'])
        monthly_est = hosts * config['hourly_rate'] * 24 * 30 * 0.6  # Rough estimate
        print(f"{name}: {int(hosts)} hosts, ~${monthly_est:,.0f}/month")
    
    print(f"\nFiles created:")
    print(f"  * {excel_file}")
    print(f"  * {report_file}")
    print(f"\nOpen the Excel file to adjust DAU and see live cost updates!")